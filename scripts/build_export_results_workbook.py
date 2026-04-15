#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXPORT_DIR = ROOT / "export_261304"
DEFAULT_BASELINE_EXPORT_DIR = ROOT / "export_261304"
DEFAULT_ACTIVE_LEARNING_DIR = ROOT / "output" / "autolabel_v1"
DEFAULT_OUTPUT_NAME = "results_overview.xlsx"

BENCHMARK_ORDER = {
    "abt-buy": 0,
    "amazon-google": 1,
    "dblp-acm": 2,
    "dblp-scholar": 3,
    "walmart-amazon": 4,
    "wdc": 5,
}
PROFILE_ORDER = {
    "official": 0,
    "small": 1,
    "small_plus20random": 2,
    "medium": 3,
    "medium_plus20random": 4,
    "large": 5,
    "all": 6,
    "all_plus20random": 7,
}
PROFILE_PARSE_ORDER = sorted(PROFILE_ORDER, key=len, reverse=True)
FAMILY_LABELS = {
    "gpt-5-mini_agent_precision": "GPT-5-mini agent_precision",
    "gpt-5-mini_agent_precision_recall": "GPT-5-mini precision->recall relabel",
    "simple_labeling": "Simple labeling",
    "three_phase_labeling_ditto_only": "Three-phase Ditto-only v1",
    "three_phase_labeling_ditto_only_v2": "Three-phase Ditto-only v2",
    "three_phase_active_learning_v2": "Three-phase Active Learning v2",
    "active_learning_v1": "Active Learning v1",
    "three_phase_labeling_ditto_only_v2_relabel_batch_gpt-5-mini_agent_precision": (
        "Three-phase v2 + batch relabel"
    ),
}
FAMILY_NOTES = {
    "gpt-5-mini_agent_precision": "Three-phase labeling run with agent_precision prompting.",
    "gpt-5-mini_agent_precision_recall": (
        "Realtime relabeling on top of the agent_precision labels with recall-oriented review."
    ),
    "simple_labeling": "Single-stage labeling baseline without the later three-phase pipeline.",
    "three_phase_labeling_ditto_only": "Early three-phase Ditto-only labeling variant.",
    "three_phase_labeling_ditto_only_v2": "Three-phase Ditto-only labeling v2 without extra relabel pass.",
    "three_phase_active_learning_v2": "Three-phase active learning v2 with rebuilt profile exports.",
    "active_learning_v1": "Legacy active learning profile sweeps from output/autolabel_v1.",
    "three_phase_labeling_ditto_only_v2_relabel_batch_gpt-5-mini_agent_precision": (
        "Three-phase v2 labels followed by a GPT-5-mini batch relabel pass."
    ),
}


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build an Excel workbook that compares exported benchmark results, baselines, "
            "and downstream Ditto training from generated labels."
        )
    )
    parser.add_argument(
        "export_dir",
        nargs="?",
        default=str(DEFAULT_EXPORT_DIR),
        help="Path to the exported result directory. Default: %(default)s",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Target .xlsx path. Default: <export_dir>/results_overview.xlsx",
    )
    parser.add_argument(
        "--baseline-export-dir",
        default=str(DEFAULT_BASELINE_EXPORT_DIR),
        help=(
            "Fallback export directory for baseline and gold-profile reference data when the "
            "current export does not include ditto_benchmark_runs/ditto_runs. Default: %(default)s"
        ),
    )
    parser.add_argument(
        "--active-learning-dir",
        default=str(DEFAULT_ACTIVE_LEARNING_DIR),
        help=(
            "Optional directory with legacy active-learning Ditto results to merge into the workbook. "
            "Default: %(default)s"
        ),
    )
    return parser.parse_args()


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text())


def _safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _family_label(family: str) -> str:
    return FAMILY_LABELS.get(family, family.replace("_", " "))


def _family_note(family: str) -> str:
    return FAMILY_NOTES.get(family, "")


def _benchmark_sort_key(value: Any) -> int:
    return BENCHMARK_ORDER.get(str(value), 99)


def _profile_sort_key(value: Any) -> int:
    return PROFILE_ORDER.get(str(value), 99)


def _extract_cost_usd(manifest: Dict[str, Any]) -> Optional[float]:
    combined = _safe_float(manifest.get("combined_labeling_cost_usd"))
    if combined is not None:
        return combined
    labeling_cost = manifest.get("labeling_cost")
    if isinstance(labeling_cost, dict):
        for key in ("total_cost_usd", "combined_cost_usd"):
            parsed = _safe_float(labeling_cost.get(key))
            if parsed is not None:
                return parsed
    return None


def _parse_profile_from_run_name(run_name: str) -> str:
    for profile in PROFILE_PARSE_ORDER:
        if re.search(rf"_{re.escape(profile)}_\d{{8}}_\d{{6}}$", run_name):
            return profile
    return "unknown"


def _parse_run_timestamp(run_name: str) -> str:
    match = re.search(r"(\d{8}_\d{6})$", run_name)
    return match.group(1) if match else ""


def _resolve_baseline_source_dirs(export_dir: Path, baseline_export_dir: Optional[Path]) -> List[Path]:
    dirs: List[Path] = [export_dir]
    if baseline_export_dir and baseline_export_dir.resolve() != export_dir.resolve():
        dirs.append(baseline_export_dir)
    return dirs


def collect_baseline_runs(
    export_dir: Path,
    baseline_export_dir: Optional[Path] = None,
) -> Tuple[pd.DataFrame, Dict[str, Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    for source_dir in _resolve_baseline_source_dirs(export_dir, baseline_export_dir):
        for summary_path in sorted((source_dir / "ditto_benchmark_runs").glob("*/summary.csv")):
            frame = pd.read_csv(summary_path)
            run_group = summary_path.parent.name
            for record in frame.to_dict("records"):
                test_f1 = _safe_float(record.get("test_f1"))
                if test_f1 is None:
                    continue
                benchmark = str(record.get("benchmark"))
                rows.append(
                    {
                        "benchmark": benchmark,
                        "baseline_run_group": run_group,
                        "baseline_test_f1": test_f1,
                        "baseline_test_precision": _safe_float(record.get("test_precision")),
                        "baseline_test_recall": _safe_float(record.get("test_recall")),
                        "baseline_test_accuracy": _safe_float(record.get("test_accuracy")),
                        "baseline_best_val_f1": _safe_float(record.get("best_val_f1")),
                        "baseline_train_rows": _safe_float(record.get("train_rows")),
                        "baseline_valid_rows": _safe_float(record.get("valid_rows")),
                        "baseline_test_rows": _safe_float(record.get("test_rows")),
                        "baseline_best_epoch": _safe_float(record.get("best_epoch")),
                        "baseline_run_dir": record.get("run_dir"),
                        "baseline_summary_csv": str(summary_path),
                    }
                )

    baseline_df = pd.DataFrame(rows)
    if baseline_df.empty:
        raise RuntimeError(
            "No baseline summary rows with test_f1 found in ditto_benchmark_runs "
            f"under {export_dir} or fallback {baseline_export_dir}."
        )

    baseline_df = baseline_df.sort_values(
        by=["benchmark", "baseline_test_f1", "baseline_run_group"],
        ascending=[True, False, False],
        kind="stable",
    )
    best_baselines = baseline_df.drop_duplicates(subset=["benchmark"], keep="first").copy()
    best_baselines["benchmark_order"] = best_baselines["benchmark"].map(_benchmark_sort_key)
    best_baselines = best_baselines.sort_values(["benchmark_order", "benchmark"]).drop(
        columns=["benchmark_order"]
    )
    baseline_index = {
        row["benchmark"]: row for row in best_baselines.to_dict("records")
    }
    return best_baselines.reset_index(drop=True), baseline_index


def collect_labeling_inventory(
    export_dir: Path,
) -> Tuple[pd.DataFrame, Dict[Tuple[str, str, str], Dict[str, Any]]]:
    rows: List[Dict[str, Any]] = []
    index: Dict[Tuple[str, str, str], Dict[str, Any]] = {}

    for manifest_path in sorted(export_dir.rglob("profile_manifest.json")):
        if "training_from_generated_labels" in manifest_path.parts:
            continue
        rel = manifest_path.relative_to(export_dir)
        family = rel.parts[0]
        manifest = _load_json(manifest_path)
        benchmark = str(manifest.get("benchmark", "")).strip()
        profiles = manifest.get("profiles") or {}
        if not benchmark or not isinstance(profiles, dict):
            continue

        label_cost_usd = _extract_cost_usd(manifest)
        runner_script = str(manifest.get("runner_script", "") or "")
        for profile, cfg in profiles.items():
            if not isinstance(cfg, dict):
                continue
            row = {
                "family": family,
                "family_label": _family_label(family),
                "family_note": _family_note(family),
                "benchmark": benchmark,
                "profile": profile,
                "all_examples": bool(cfg.get("all_examples", False)),
                "base_profile": cfg.get("base_profile"),
                "random_fraction": _safe_float(cfg.get("random_fraction")),
                "random_additions": _safe_float(cfg.get("random_additions")),
                "shared_random_model": cfg.get("shared_random_model"),
                "target_total": _safe_float(cfg.get("target_total")),
                "target_pos": _safe_float(cfg.get("target_pos")),
                "target_neg": _safe_float(cfg.get("target_neg")),
                "actual_total": _safe_float(cfg.get("actual_total")),
                "actual_pos": _safe_float(cfg.get("actual_pos")),
                "actual_neg": _safe_float(cfg.get("actual_neg")),
                "label_cost_usd": label_cost_usd,
                "runner_script": runner_script,
                "manifest_path": str(manifest_path),
                "run_dir": manifest.get("run_dir"),
            }
            rows.append(row)
            index[(family, benchmark, profile)] = row

    inventory_df = pd.DataFrame(rows)
    if inventory_df.empty:
        return inventory_df, index

    inventory_df["benchmark_order"] = inventory_df["benchmark"].map(_benchmark_sort_key)
    inventory_df["profile_order"] = inventory_df["profile"].map(_profile_sort_key)
    inventory_df = inventory_df.sort_values(
        ["benchmark_order", "benchmark", "family_label", "profile_order", "profile"],
        kind="stable",
    ).drop(columns=["benchmark_order", "profile_order"])
    return inventory_df.reset_index(drop=True), index


def collect_generated_training_runs(
    export_dir: Path,
    manifest_index: Dict[Tuple[str, str, str], Dict[str, Any]],
    baseline_index: Dict[str, Dict[str, Any]],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    rows: List[Dict[str, Any]] = []
    nested_root = export_dir / "training_from_generated_labels"
    if nested_root.exists():
        summary_entries = [
            (summary_path, summary_path.relative_to(nested_root).parts[0])
            for summary_path in sorted(nested_root.rglob("summary.csv"))
        ]
    else:
        summary_entries = []
        for family_dir in sorted(path for path in export_dir.iterdir() if path.is_dir()):
            if family_dir.name in {"ditto_benchmark_runs", "ditto_runs"}:
                continue
            family_summaries = sorted(family_dir.rglob("summary.csv"))
            if not family_summaries:
                continue
            summary_entries.extend((summary_path, family_dir.name) for summary_path in family_summaries)

    if not summary_entries:
        return pd.DataFrame(), pd.DataFrame()

    for summary_path, family in summary_entries:
        run_name = summary_path.parent.name
        profile = _parse_profile_from_run_name(run_name)
        run_timestamp = _parse_run_timestamp(run_name)
        frame = pd.read_csv(summary_path)
        if frame.empty:
            continue

        for record in frame.to_dict("records"):
            benchmark = str(record.get("benchmark"))
            manifest_row = manifest_index.get((family, benchmark, profile), {})
            baseline_row = baseline_index.get(benchmark, {})
            test_f1 = _safe_float(record.get("test_f1"))
            baseline_f1 = _safe_float(baseline_row.get("baseline_test_f1"))
            train_rows = _safe_float(record.get("train_rows"))
            baseline_train_rows = _safe_float(baseline_row.get("baseline_train_rows"))

            rows.append(
                {
                    "benchmark": benchmark,
                    "family": family,
                    "family_label": _family_label(family),
                    "family_note": _family_note(family),
                    "profile": profile,
                    "run_name": run_name,
                    "run_timestamp": run_timestamp,
                    "summary_csv": str(summary_path),
                    "status": record.get("status"),
                    "test_f1": test_f1,
                    "test_precision": _safe_float(record.get("test_precision")),
                    "test_recall": _safe_float(record.get("test_recall")),
                    "test_accuracy": _safe_float(record.get("test_accuracy")),
                    "best_val_f1": _safe_float(record.get("best_val_f1")),
                    "best_epoch": _safe_float(record.get("best_epoch")),
                    "train_rows": train_rows,
                    "valid_rows": _safe_float(record.get("valid_rows")),
                    "test_rows": _safe_float(record.get("test_rows")),
                    "run_dir": record.get("run_dir"),
                    "error": record.get("error"),
                    "baseline_test_f1": baseline_f1,
                    "baseline_train_rows": baseline_train_rows,
                    "delta_f1_vs_baseline": (
                        test_f1 - baseline_f1 if test_f1 is not None and baseline_f1 is not None else None
                    ),
                    "delta_precision_vs_baseline": (
                        _safe_float(record.get("test_precision"))
                        - _safe_float(baseline_row.get("baseline_test_precision"))
                        if _safe_float(record.get("test_precision")) is not None
                        and _safe_float(baseline_row.get("baseline_test_precision")) is not None
                        else None
                    ),
                    "delta_recall_vs_baseline": (
                        _safe_float(record.get("test_recall"))
                        - _safe_float(baseline_row.get("baseline_test_recall"))
                        if _safe_float(record.get("test_recall")) is not None
                        and _safe_float(baseline_row.get("baseline_test_recall")) is not None
                        else None
                    ),
                    "label_rows_vs_gold_train": (
                        train_rows / baseline_train_rows
                        if train_rows is not None and baseline_train_rows not in (None, 0)
                        else None
                    ),
                    "target_total": manifest_row.get("target_total"),
                    "target_pos": manifest_row.get("target_pos"),
                    "target_neg": manifest_row.get("target_neg"),
                    "actual_total": manifest_row.get("actual_total"),
                    "actual_pos": manifest_row.get("actual_pos"),
                    "actual_neg": manifest_row.get("actual_neg"),
                    "label_cost_usd": manifest_row.get("label_cost_usd"),
                    "runner_script": manifest_row.get("runner_script"),
                    "manifest_path": manifest_row.get("manifest_path"),
                }
            )

    all_runs_df = pd.DataFrame(rows)
    if all_runs_df.empty:
        return all_runs_df, pd.DataFrame()

    all_runs_df["benchmark_order"] = all_runs_df["benchmark"].map(_benchmark_sort_key)
    all_runs_df["profile_order"] = all_runs_df["profile"].map(_profile_sort_key)

    group_cols = ["benchmark", "family", "family_label", "profile"]
    run_counts = (
        all_runs_df.groupby(group_cols, dropna=False)
        .agg(
            total_runs=("run_name", "size"),
            successful_runs=("status", lambda s: int((pd.Series(s) == "ok").sum())),
            failed_runs=("status", lambda s: int((pd.Series(s) != "ok").sum())),
        )
        .reset_index()
    )

    successful_runs = all_runs_df[
        (all_runs_df["status"] == "ok") & all_runs_df["test_f1"].notna()
    ].copy()
    if successful_runs.empty:
        return all_runs_df.reset_index(drop=True), pd.DataFrame()

    successful_runs = successful_runs.sort_values(
        by=[
            "benchmark_order",
            "benchmark",
            "family_label",
            "profile_order",
            "test_f1",
            "run_timestamp",
        ],
        ascending=[True, True, True, True, False, False],
        kind="stable",
    )
    best_config_df = successful_runs.drop_duplicates(
        subset=["benchmark", "family", "profile"], keep="first"
    )
    mean_metrics = (
        successful_runs.groupby(group_cols, dropna=False)
        .agg(
            mean_test_f1=("test_f1", "mean"),
            mean_test_precision=("test_precision", "mean"),
            mean_test_recall=("test_recall", "mean"),
            mean_test_accuracy=("test_accuracy", "mean"),
        )
        .reset_index()
    )

    best_config_df = best_config_df.merge(run_counts, on=group_cols, how="left").merge(
        mean_metrics, on=group_cols, how="left"
    )
    best_config_df = best_config_df.sort_values(
        by=["benchmark_order", "benchmark", "test_f1", "profile_order"],
        ascending=[True, True, False, True],
        kind="stable",
    )
    return (
        all_runs_df.sort_values(
            ["benchmark_order", "benchmark", "family_label", "profile_order", "run_timestamp"],
            kind="stable",
        ).drop(columns=["benchmark_order", "profile_order"]).reset_index(drop=True),
        best_config_df.drop(columns=["benchmark_order", "profile_order"]).reset_index(drop=True),
    )


def _parse_timestamp_from_name(value: str) -> str:
    match = re.search(r"(\d{8}_\d{6})$", str(value))
    return match.group(1) if match else ""


def collect_legacy_active_learning_runs(
    active_learning_dir: Optional[Path],
    baseline_index: Dict[str, Dict[str, Any]],
) -> pd.DataFrame:
    if active_learning_dir is None or not active_learning_dir.exists():
        return pd.DataFrame()

    rows: List[Dict[str, Any]] = []
    family = "active_learning_v1"

    for metrics_path in sorted(active_learning_dir.glob("*/metrics.tsv")):
        frame = pd.read_csv(metrics_path, sep="\t")
        for record in frame.to_dict("records"):
            benchmark = str(record.get("benchmark"))
            profile = str(record.get("profile"))
            run_dir = str(record.get("run_dir", "") or "")
            run_name = Path(run_dir).name if run_dir else metrics_path.parent.name
            baseline_row = baseline_index.get(benchmark, {})
            test_f1 = _safe_float(record.get("test_f1"))
            rows.append(
                {
                    "benchmark": benchmark,
                    "family": family,
                    "family_label": _family_label(family),
                    "family_note": _family_note(family),
                    "profile": profile,
                    "run_name": run_name,
                    "run_timestamp": _parse_timestamp_from_name(run_name),
                    "summary_csv": str(metrics_path),
                    "status": "ok",
                    "test_f1": test_f1,
                    "test_precision": _safe_float(record.get("test_precision")),
                    "test_recall": _safe_float(record.get("test_recall")),
                    "test_accuracy": _safe_float(record.get("test_accuracy")),
                    "best_val_f1": _safe_float(record.get("best_val_f1")),
                    "best_epoch": None,
                    "train_rows": None,
                    "valid_rows": None,
                    "test_rows": None,
                    "run_dir": run_dir,
                    "error": None,
                    "baseline_test_f1": _safe_float(baseline_row.get("baseline_test_f1")),
                    "baseline_train_rows": _safe_float(baseline_row.get("baseline_train_rows")),
                    "delta_f1_vs_baseline": (
                        test_f1 - _safe_float(baseline_row.get("baseline_test_f1"))
                        if test_f1 is not None and _safe_float(baseline_row.get("baseline_test_f1")) is not None
                        else None
                    ),
                    "delta_precision_vs_baseline": (
                        _safe_float(record.get("test_precision"))
                        - _safe_float(baseline_row.get("baseline_test_precision"))
                        if _safe_float(record.get("test_precision")) is not None
                        and _safe_float(baseline_row.get("baseline_test_precision")) is not None
                        else None
                    ),
                    "delta_recall_vs_baseline": (
                        _safe_float(record.get("test_recall"))
                        - _safe_float(baseline_row.get("baseline_test_recall"))
                        if _safe_float(record.get("test_recall")) is not None
                        and _safe_float(baseline_row.get("baseline_test_recall")) is not None
                        else None
                    ),
                    "label_rows_vs_gold_train": None,
                    "target_total": None,
                    "target_pos": None,
                    "target_neg": None,
                    "actual_total": None,
                    "actual_pos": None,
                    "actual_neg": None,
                    "label_cost_usd": None,
                    "runner_script": "legacy active learning",
                    "manifest_path": str(metrics_path),
                }
            )

    abt_root = active_learning_dir / "abt_buy_profiles_abt_buy_local_test_173424"
    for metrics_json in sorted(abt_root.rglob("metrics.json")):
        profile = metrics_json.parents[1].name
        payload = _load_json(metrics_json)
        test = payload.get("test") or {}
        benchmark = "abt-buy"
        baseline_row = baseline_index.get(benchmark, {})
        test_f1 = _safe_float(test.get("f1"))
        run_name = metrics_json.parent.name
        rows.append(
            {
                "benchmark": benchmark,
                "family": family,
                "family_label": _family_label(family),
                "family_note": _family_note(family),
                "profile": profile,
                "run_name": run_name,
                "run_timestamp": _parse_timestamp_from_name(run_name),
                "summary_csv": str(metrics_json),
                "status": "ok",
                "test_f1": test_f1,
                "test_precision": _safe_float(test.get("precision")),
                "test_recall": _safe_float(test.get("recall")),
                "test_accuracy": _safe_float(test.get("accuracy")),
                "best_val_f1": _safe_float(payload.get("best_val_f1")),
                "best_epoch": None,
                "train_rows": None,
                "valid_rows": None,
                "test_rows": None,
                "run_dir": str(metrics_json.parent),
                "error": None,
                "baseline_test_f1": _safe_float(baseline_row.get("baseline_test_f1")),
                "baseline_train_rows": _safe_float(baseline_row.get("baseline_train_rows")),
                "delta_f1_vs_baseline": (
                    test_f1 - _safe_float(baseline_row.get("baseline_test_f1"))
                    if test_f1 is not None and _safe_float(baseline_row.get("baseline_test_f1")) is not None
                    else None
                ),
                "delta_precision_vs_baseline": (
                    _safe_float(test.get("precision"))
                    - _safe_float(baseline_row.get("baseline_test_precision"))
                    if _safe_float(test.get("precision")) is not None
                    and _safe_float(baseline_row.get("baseline_test_precision")) is not None
                    else None
                ),
                "delta_recall_vs_baseline": (
                    _safe_float(test.get("recall"))
                    - _safe_float(baseline_row.get("baseline_test_recall"))
                    if _safe_float(test.get("recall")) is not None
                    and _safe_float(baseline_row.get("baseline_test_recall")) is not None
                    else None
                ),
                "label_rows_vs_gold_train": None,
                "target_total": None,
                "target_pos": None,
                "target_neg": None,
                "actual_total": None,
                "actual_pos": None,
                "actual_neg": None,
                "label_cost_usd": None,
                "runner_script": "legacy active learning",
                "manifest_path": str(metrics_json),
            }
        )

    wdc_metrics = active_learning_dir / "wdc" / "run_20260220_235713" / "metrics.json"
    if wdc_metrics.exists():
        payload = _load_json(wdc_metrics)
        test = payload.get("test") or {}
        benchmark = "wdc"
        baseline_row = baseline_index.get(benchmark, {})
        test_f1 = _safe_float(test.get("f1"))
        rows.append(
            {
                "benchmark": benchmark,
                "family": family,
                "family_label": _family_label(family),
                "family_note": _family_note(family),
                "profile": "legacy",
                "run_name": wdc_metrics.parent.name,
                "run_timestamp": _parse_timestamp_from_name(wdc_metrics.parent.name),
                "summary_csv": str(wdc_metrics),
                "status": "ok",
                "test_f1": test_f1,
                "test_precision": _safe_float(test.get("precision")),
                "test_recall": _safe_float(test.get("recall")),
                "test_accuracy": _safe_float(test.get("accuracy")),
                "best_val_f1": _safe_float(payload.get("best_val_f1")),
                "best_epoch": None,
                "train_rows": None,
                "valid_rows": None,
                "test_rows": None,
                "run_dir": str(wdc_metrics.parent),
                "error": None,
                "baseline_test_f1": _safe_float(baseline_row.get("baseline_test_f1")),
                "baseline_train_rows": _safe_float(baseline_row.get("baseline_train_rows")),
                "delta_f1_vs_baseline": (
                    test_f1 - _safe_float(baseline_row.get("baseline_test_f1"))
                    if test_f1 is not None and _safe_float(baseline_row.get("baseline_test_f1")) is not None
                    else None
                ),
                "delta_precision_vs_baseline": (
                    _safe_float(test.get("precision"))
                    - _safe_float(baseline_row.get("baseline_test_precision"))
                    if _safe_float(test.get("precision")) is not None
                    and _safe_float(baseline_row.get("baseline_test_precision")) is not None
                    else None
                ),
                "delta_recall_vs_baseline": (
                    _safe_float(test.get("recall"))
                    - _safe_float(baseline_row.get("baseline_test_recall"))
                    if _safe_float(test.get("recall")) is not None
                    and _safe_float(baseline_row.get("baseline_test_recall")) is not None
                    else None
                ),
                "label_rows_vs_gold_train": None,
                "target_total": None,
                "target_pos": None,
                "target_neg": None,
                "actual_total": None,
                "actual_pos": None,
                "actual_neg": None,
                "label_cost_usd": None,
                "runner_script": "legacy active learning",
                "manifest_path": str(wdc_metrics),
            }
        )

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    df["benchmark_order"] = df["benchmark"].map(_benchmark_sort_key)
    df["profile_order"] = df["profile"].map(_profile_sort_key)
    df = df.sort_values(
        ["benchmark_order", "benchmark", "family_label", "profile_order", "run_timestamp"],
        kind="stable",
    ).drop(columns=["benchmark_order", "profile_order"])
    return df.reset_index(drop=True)


def collect_gold_profile_reference(
    export_dir: Path,
    baseline_index: Dict[str, Dict[str, Any]],
    baseline_export_dir: Optional[Path] = None,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    seen_metrics: set[str] = set()
    for source_dir in _resolve_baseline_source_dirs(export_dir, baseline_export_dir):
        for metrics_path in sorted((source_dir / "ditto_runs").glob("*/metrics.tsv")):
            if str(metrics_path.resolve()) in seen_metrics:
                continue
            seen_metrics.add(str(metrics_path.resolve()))
            experiment = metrics_path.parent.name
            frame = pd.read_csv(metrics_path, sep="\t")
            for record in frame.to_dict("records"):
                benchmark = str(record.get("benchmark"))
                baseline_row = baseline_index.get(benchmark, {})
                test_f1 = _safe_float(record.get("test_f1"))
                baseline_f1 = _safe_float(baseline_row.get("baseline_test_f1"))
                rows.append(
                    {
                        "benchmark": benchmark,
                        "experiment": experiment,
                        "profile": record.get("profile"),
                        "test_f1": test_f1,
                        "test_precision": _safe_float(record.get("test_precision")),
                        "test_recall": _safe_float(record.get("test_recall")),
                        "test_accuracy": _safe_float(record.get("test_accuracy")),
                        "baseline_test_f1": baseline_f1,
                        "delta_f1_vs_baseline": (
                            test_f1 - baseline_f1 if test_f1 is not None and baseline_f1 is not None else None
                        ),
                        "run_dir": record.get("run_dir"),
                    }
                )

    gold_df = pd.DataFrame(rows)
    if gold_df.empty:
        return gold_df

    gold_df["benchmark_order"] = gold_df["benchmark"].map(_benchmark_sort_key)
    gold_df["profile_order"] = gold_df["profile"].map(_profile_sort_key)
    gold_df = gold_df.sort_values(
        ["benchmark_order", "benchmark", "profile_order", "profile", "experiment"],
        kind="stable",
    ).drop(columns=["benchmark_order", "profile_order"])
    return gold_df.reset_index(drop=True)


def build_overview(
    baseline_df: pd.DataFrame,
    best_config_df: pd.DataFrame,
    all_runs_df: pd.DataFrame,
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    best_by_benchmark = {
        row["benchmark"]: row
        for row in best_config_df.sort_values(
            ["benchmark", "test_f1", "run_timestamp"], ascending=[True, False, False]
        ).drop_duplicates("benchmark", keep="first").to_dict("records")
    }
    run_counts = (
        all_runs_df.groupby("benchmark")
        .agg(
            total_runs=("run_name", "size"),
            successful_runs=("status", lambda s: int((pd.Series(s) == "ok").sum())),
            failed_runs=("status", lambda s: int((pd.Series(s) != "ok").sum())),
        )
        .to_dict("index")
        if not all_runs_df.empty
        else {}
    )
    config_counts = (
        best_config_df.groupby("benchmark").size().to_dict() if not best_config_df.empty else {}
    )

    for record in baseline_df.to_dict("records"):
        benchmark = record["benchmark"]
        best = best_by_benchmark.get(benchmark, {})
        counts = run_counts.get(benchmark, {})
        rows.append(
            {
                "benchmark": benchmark,
                "baseline_test_f1": record.get("baseline_test_f1"),
                "baseline_test_precision": record.get("baseline_test_precision"),
                "baseline_test_recall": record.get("baseline_test_recall"),
                "best_configuration": best.get("family_label"),
                "best_profile": best.get("profile"),
                "best_test_f1": best.get("test_f1"),
                "delta_f1_vs_baseline": best.get("delta_f1_vs_baseline"),
                "best_test_precision": best.get("test_precision"),
                "best_test_recall": best.get("test_recall"),
                "best_label_rows": best.get("actual_total"),
                "best_label_pos": best.get("actual_pos"),
                "best_label_neg": best.get("actual_neg"),
                "best_label_cost_usd": best.get("label_cost_usd"),
                "successful_configurations": config_counts.get(benchmark, 0),
                "successful_runs": counts.get("successful_runs", 0),
                "failed_runs": counts.get("failed_runs", 0),
            }
        )

    overview_df = pd.DataFrame(rows)
    overview_df["benchmark_order"] = overview_df["benchmark"].map(_benchmark_sort_key)
    overview_df = overview_df.sort_values(["benchmark_order", "benchmark"]).drop(
        columns=["benchmark_order"]
    )
    return overview_df.reset_index(drop=True)


def build_config_summary(best_config_df: pd.DataFrame) -> pd.DataFrame:
    if best_config_df.empty:
        return pd.DataFrame()

    summary = (
        best_config_df.groupby(["family", "family_label", "profile", "family_note"], dropna=False)
        .agg(
            benchmarks_tested=("benchmark", "nunique"),
            benchmark_list=("benchmark", lambda s: ", ".join(sorted(set(str(v) for v in s)))),
            best_test_f1=("test_f1", "max"),
            avg_test_f1=("test_f1", "mean"),
            avg_delta_f1_vs_baseline=("delta_f1_vs_baseline", "mean"),
            max_delta_f1_vs_baseline=("delta_f1_vs_baseline", "max"),
            min_delta_f1_vs_baseline=("delta_f1_vs_baseline", "min"),
            configs_beating_baseline=("delta_f1_vs_baseline", lambda s: int((pd.Series(s) > 0).sum())),
            avg_label_rows=("actual_total", "mean"),
            avg_label_pos=("actual_pos", "mean"),
            avg_label_neg=("actual_neg", "mean"),
            total_successful_runs=("successful_runs", "sum"),
            total_failed_runs=("failed_runs", "sum"),
        )
        .reset_index()
    )
    summary["profile_order"] = summary["profile"].map(_profile_sort_key)
    summary = summary.sort_values(
        ["avg_delta_f1_vs_baseline", "benchmarks_tested", "profile_order"],
        ascending=[False, False, True],
        kind="stable",
    ).drop(columns=["profile_order"])
    return summary.reset_index(drop=True)


def build_dataset_tabs(
    baseline_df: pd.DataFrame,
    best_config_df: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    tabs: Dict[str, pd.DataFrame] = {}
    baseline_index = {
        row["benchmark"]: row for row in baseline_df.to_dict("records")
    }

    for benchmark in baseline_df["benchmark"].tolist():
        baseline = baseline_index[benchmark]
        dataset_rows = best_config_df[best_config_df["benchmark"] == benchmark].copy()
        dataset_rows = dataset_rows.sort_values(
            ["test_f1", "delta_f1_vs_baseline", "profile"],
            ascending=[False, False, True],
            kind="stable",
        )

        baseline_row = {
            "rank": "",
            "configuration": "Gold baseline",
            "profile": "official",
            "successful_runs": 1,
            "failed_runs": 0,
            "test_f1": baseline.get("baseline_test_f1"),
            "delta_f1_vs_baseline": 0.0,
            "test_precision": baseline.get("baseline_test_precision"),
            "test_recall": baseline.get("baseline_test_recall"),
            "test_accuracy": baseline.get("baseline_test_accuracy"),
            "train_rows": baseline.get("baseline_train_rows"),
            "label_rows_vs_gold_train": 1.0,
            "actual_total": baseline.get("baseline_train_rows"),
            "actual_pos": None,
            "actual_neg": None,
            "label_cost_usd": None,
            "runner_script": "Gold labels",
            "best_run_name": baseline.get("baseline_run_group"),
            "best_run_dir": baseline.get("baseline_run_dir"),
            "family_note": "Official Ditto benchmark training on gold labels.",
        }
        records = [baseline_row]
        for rank, row in enumerate(dataset_rows.to_dict("records"), start=1):
            records.append(
                {
                    "rank": rank,
                    "configuration": row.get("family_label"),
                    "profile": row.get("profile"),
                    "successful_runs": row.get("successful_runs"),
                    "failed_runs": row.get("failed_runs"),
                    "test_f1": row.get("test_f1"),
                    "delta_f1_vs_baseline": row.get("delta_f1_vs_baseline"),
                    "test_precision": row.get("test_precision"),
                    "test_recall": row.get("test_recall"),
                    "test_accuracy": row.get("test_accuracy"),
                    "train_rows": row.get("train_rows"),
                    "label_rows_vs_gold_train": row.get("label_rows_vs_gold_train"),
                    "actual_total": row.get("actual_total"),
                    "actual_pos": row.get("actual_pos"),
                    "actual_neg": row.get("actual_neg"),
                    "label_cost_usd": row.get("label_cost_usd"),
                    "runner_script": row.get("runner_script"),
                    "best_run_name": row.get("run_name"),
                    "best_run_dir": row.get("run_dir"),
                    "family_note": row.get("family_note"),
                }
            )
        tabs[benchmark] = pd.DataFrame(records)

    return tabs


def build_failures_tab(all_runs_df: pd.DataFrame) -> pd.DataFrame:
    if all_runs_df.empty:
        return pd.DataFrame()
    failures = all_runs_df[all_runs_df["status"] != "ok"].copy()
    if failures.empty:
        return failures
    failures = failures[
        [
            "benchmark",
            "family_label",
            "profile",
            "run_name",
            "status",
            "error",
            "summary_csv",
        ]
    ].rename(columns={"family_label": "configuration"})
    failures["benchmark_order"] = failures["benchmark"].map(_benchmark_sort_key)
    failures["profile_order"] = failures["profile"].map(_profile_sort_key)
    failures = failures.sort_values(
        ["benchmark_order", "benchmark", "configuration", "profile_order", "run_name"],
        kind="stable",
    ).drop(columns=["benchmark_order", "profile_order"])
    return failures.reset_index(drop=True)


def _apply_workbook_formatting(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    positive_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    negative_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")

    percent_cols = {
        "baseline_test_f1",
        "baseline_test_precision",
        "baseline_test_recall",
        "best_test_f1",
        "best_test_precision",
        "best_test_recall",
        "test_f1",
        "test_precision",
        "test_recall",
        "test_accuracy",
        "best_val_f1",
        "mean_test_f1",
        "mean_test_precision",
        "mean_test_recall",
        "mean_test_accuracy",
        "avg_test_f1",
        "best_test_f1",
        "label_rows_vs_gold_train",
    }
    delta_cols = {
        "delta_f1_vs_baseline",
        "delta_precision_vs_baseline",
        "delta_recall_vs_baseline",
        "avg_delta_f1_vs_baseline",
        "max_delta_f1_vs_baseline",
        "min_delta_f1_vs_baseline",
    }
    currency_cols = {"label_cost_usd", "best_label_cost_usd"}

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        headers = [cell.value for cell in sheet[1]]
        header_map = {idx + 1: str(value) for idx, value in enumerate(headers) if value is not None}

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for col_idx, header in header_map.items():
            max_len = len(header)
            for cell in sheet[get_column_letter(col_idx)]:
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
                if cell.row == 1:
                    continue
                if header in percent_cols and isinstance(value, (int, float)):
                    cell.number_format = "0.0%"
                elif header in delta_cols and isinstance(value, (int, float)):
                    cell.number_format = "+0.0%;-0.0%;0.0%"
                    if value > 0:
                        cell.fill = positive_fill
                    elif value < 0:
                        cell.fill = negative_fill
                elif header in currency_cols and isinstance(value, (int, float)):
                    cell.number_format = "$0.000"
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 55)

    workbook.save(path)


def write_workbook(
    output_path: Path,
    overview_df: pd.DataFrame,
    config_summary_df: pd.DataFrame,
    labeling_inventory_df: pd.DataFrame,
    gold_profile_df: pd.DataFrame,
    failures_df: pd.DataFrame,
    dataset_tabs: Dict[str, pd.DataFrame],
    all_runs_df: pd.DataFrame,
    best_config_df: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Overview", index=False)
        config_summary_df.to_excel(writer, sheet_name="Config Summary", index=False)
        labeling_inventory_df.to_excel(writer, sheet_name="Labeling Inventory", index=False)
        gold_profile_df.to_excel(writer, sheet_name="Gold Profiles", index=False)
        failures_df.to_excel(writer, sheet_name="Failures", index=False)
        best_config_df.to_excel(writer, sheet_name="All Best Configs", index=False)
        all_runs_df.to_excel(writer, sheet_name="All Runs", index=False)
        for benchmark, frame in dataset_tabs.items():
            frame.to_excel(writer, sheet_name=benchmark[:31], index=False)

    _apply_workbook_formatting(output_path)


def main() -> None:
    args = _parse_args()
    export_dir = Path(args.export_dir).expanduser().resolve()
    if not export_dir.exists():
        raise FileNotFoundError(f"Export directory does not exist: {export_dir}")

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else export_dir / DEFAULT_OUTPUT_NAME
    )
    baseline_export_dir = Path(args.baseline_export_dir).expanduser().resolve() if args.baseline_export_dir else None
    active_learning_dir = Path(args.active_learning_dir).expanduser().resolve() if args.active_learning_dir else None

    baseline_df, baseline_index = collect_baseline_runs(export_dir, baseline_export_dir=baseline_export_dir)
    labeling_inventory_df, manifest_index = collect_labeling_inventory(export_dir)
    all_runs_df, best_config_df = collect_generated_training_runs(
        export_dir=export_dir,
        manifest_index=manifest_index,
        baseline_index=baseline_index,
    )
    legacy_active_learning_df = collect_legacy_active_learning_runs(
        active_learning_dir=active_learning_dir,
        baseline_index=baseline_index,
    )
    if not legacy_active_learning_df.empty:
        if all_runs_df.empty:
            all_runs_df = legacy_active_learning_df.copy()
        else:
            all_runs_df = pd.concat([all_runs_df, legacy_active_learning_df], ignore_index=True, sort=False)
        successful_legacy = legacy_active_learning_df[
            (legacy_active_learning_df["status"] == "ok") & legacy_active_learning_df["test_f1"].notna()
        ].copy()
        if not successful_legacy.empty:
            group_cols = ["benchmark", "family", "family_label", "profile"]
            run_counts = (
                successful_legacy.groupby(group_cols, dropna=False)
                .agg(
                    total_runs=("run_name", "size"),
                    successful_runs=("status", lambda s: int((pd.Series(s) == "ok").sum())),
                    failed_runs=("status", lambda s: int((pd.Series(s) != "ok").sum())),
                )
                .reset_index()
            )
            mean_metrics = (
                successful_legacy.groupby(group_cols, dropna=False)
                .agg(
                    mean_test_f1=("test_f1", "mean"),
                    mean_test_precision=("test_precision", "mean"),
                    mean_test_recall=("test_recall", "mean"),
                    mean_test_accuracy=("test_accuracy", "mean"),
                )
                .reset_index()
            )
            legacy_best = successful_legacy.sort_values(
                by=["benchmark", "family_label", "profile", "test_f1", "run_timestamp"],
                ascending=[True, True, True, False, False],
                kind="stable",
            ).drop_duplicates(subset=["benchmark", "family", "profile"], keep="first")
            legacy_best = legacy_best.merge(run_counts, on=group_cols, how="left").merge(
                mean_metrics, on=group_cols, how="left"
            )
            if best_config_df.empty:
                best_config_df = legacy_best.copy()
            else:
                best_config_df = pd.concat([best_config_df, legacy_best], ignore_index=True, sort=False)

    if not all_runs_df.empty:
        all_runs_df["benchmark_order"] = all_runs_df["benchmark"].map(_benchmark_sort_key)
        all_runs_df["profile_order"] = all_runs_df["profile"].map(_profile_sort_key)
        all_runs_df = all_runs_df.sort_values(
            ["benchmark_order", "benchmark", "family_label", "profile_order", "run_timestamp"],
            kind="stable",
        ).drop(columns=["benchmark_order", "profile_order"]).reset_index(drop=True)
    if not best_config_df.empty:
        best_config_df["benchmark_order"] = best_config_df["benchmark"].map(_benchmark_sort_key)
        best_config_df["profile_order"] = best_config_df["profile"].map(_profile_sort_key)
        best_config_df = best_config_df.sort_values(
            ["benchmark_order", "benchmark", "test_f1", "profile_order"],
            ascending=[True, True, False, True],
            kind="stable",
        ).drop(columns=["benchmark_order", "profile_order"]).reset_index(drop=True)

    gold_profile_df = collect_gold_profile_reference(
        export_dir,
        baseline_index,
        baseline_export_dir=baseline_export_dir,
    )
    overview_df = build_overview(baseline_df, best_config_df, all_runs_df)
    config_summary_df = build_config_summary(best_config_df)
    failures_df = build_failures_tab(all_runs_df)
    dataset_tabs = build_dataset_tabs(baseline_df, best_config_df)

    write_workbook(
        output_path=output_path,
        overview_df=overview_df,
        config_summary_df=config_summary_df,
        labeling_inventory_df=labeling_inventory_df,
        gold_profile_df=gold_profile_df,
        failures_df=failures_df,
        dataset_tabs=dataset_tabs,
        all_runs_df=all_runs_df,
        best_config_df=best_config_df,
    )

    print(output_path)


if __name__ == "__main__":
    main()
