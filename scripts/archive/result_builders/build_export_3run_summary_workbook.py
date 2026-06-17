#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[3]
DEFAULT_OUTPUT_NAME = "results_3run_summary.xlsx"
TRAINING_ROOT_NAME = "training_from_generated_labels_3runs"

BENCHMARK_ORDER = {
    "abt-buy": 0,
    "amazon-google": 1,
    "dblp-acm": 2,
    "dblp-scholar": 3,
    "walmart-amazon": 4,
    "wdc": 5,
}
PROFILE_ORDER = {
    "official": 0,
    "small": 1,
    "small_plus20random": 2,
    "medium": 3,
    "medium_plus20random": 4,
    "large": 5,
    "large_plus20random": 6,
    "all": 7,
    "all_plus20random": 8,
}
PROFILE_PARSE_ORDER = sorted(PROFILE_ORDER, key=len, reverse=True)
FAMILY_LABELS = {
    "ditto_baseline": "Ditto baseline",
    "simple_active_learning_labeling": "Simple active learning",
    "three_phase_active_learning_v2": "Three-phase active learning v2",
    "three_phase_labeling_ditto_only_v2": "Three-phase Ditto-only v2",
    "three_phase_labeling_ditto_only_v2_drop_changed": "Three-phase v2 drop_changed",
    "three_phase_labeling_ditto_only_v2_closure_bridge_drop": "Three-phase v2 closure_bridge_drop",
    "three_phase_labeling_ditto_only_v2_closure_bridge_relabel_changed_drop": (
        "Three-phase v2 closure_bridge_relabel_changed_drop"
    ),
    "three_phase_labeling_ditto_only_v2_closure_bridge_or_relabel_changed_drop": (
        "Three-phase v2 closure_bridge_or_relabel_changed_drop"
    ),
    "three_phase_labeling_ditto_only_v2_relabel_batch_gpt-5-mini_agent_precision": (
        "Three-phase v2 batch relabel"
    ),
    "seed_round_only_profiles": "Seed-round-only profiles",
}


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build an Excel workbook that summarizes 3-run Ditto exports with mean and std F1."
    )
    parser.add_argument("export_dir", help="Path to the export directory.")
    parser.add_argument(
        "--output",
        default=None,
        help=f"Target .xlsx path. Default: <export_dir>/{DEFAULT_OUTPUT_NAME}",
    )
    return parser.parse_args()


def _safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _family_label(family: str) -> str:
    return FAMILY_LABELS.get(family, family.replace("_", " "))


def _parse_profile(run_name: str, family: str) -> str:
    if family == "ditto_baseline":
        return "official"
    for profile in PROFILE_PARSE_ORDER:
        if re.search(rf"_{re.escape(profile)}(?:_r\d+)?_\d{{8}}_\d{{6}}$", run_name):
            return profile
    return "unknown"


def _parse_repeat(run_name: str) -> Optional[int]:
    match = re.search(r"_r(\d+)(?:_|$)", run_name)
    if not match:
        return None
    return int(match.group(1))


def _parse_seed(run_name: str) -> Optional[int]:
    match = re.search(r"_seed(\d+)(?:_|$)", run_name)
    if not match:
        return None
    return int(match.group(1))


def _benchmark_sort_key(value: Any) -> int:
    return BENCHMARK_ORDER.get(str(value), 99)


def _profile_sort_key(value: Any) -> int:
    return PROFILE_ORDER.get(str(value), 99)


def collect_runs(export_dir: Path) -> pd.DataFrame:
    training_root = export_dir / TRAINING_ROOT_NAME
    if not training_root.exists():
        raise FileNotFoundError(f"Missing training root: {training_root}")

    rows: List[Dict[str, Any]] = []
    for summary_path in sorted(training_root.rglob("summary.csv")):
        rel = summary_path.relative_to(training_root)
        if len(rel.parts) < 3:
            continue
        family = rel.parts[0]
        run_name = rel.parts[1]
        profile = _parse_profile(run_name, family)
        repeat_idx = _parse_repeat(run_name)
        seed = _parse_seed(run_name)

        frame = pd.read_csv(summary_path)
        if frame.empty:
            continue

        for record in frame.to_dict("records"):
            test_f1 = _safe_float(record.get("test_f1"))
            if str(record.get("status", "")).strip().lower() != "ok" or test_f1 is None:
                continue
            rows.append(
                {
                    "family": family,
                    "family_label": _family_label(family),
                    "run_name": run_name,
                    "profile": profile,
                    "repeat_idx": repeat_idx,
                    "seed": seed,
                    "benchmark": str(record.get("benchmark")),
                    "test_f1": test_f1,
                    "test_precision": _safe_float(record.get("test_precision")),
                    "test_recall": _safe_float(record.get("test_recall")),
                    "test_accuracy": _safe_float(record.get("test_accuracy")),
                    "best_val_f1": _safe_float(record.get("best_val_f1")),
                    "train_rows": _safe_float(record.get("train_rows")),
                    "valid_rows": _safe_float(record.get("valid_rows")),
                    "test_rows": _safe_float(record.get("test_rows")),
                    "summary_csv": str(summary_path),
                    "run_dir": record.get("run_dir"),
                }
            )

    out = pd.DataFrame(rows)
    if out.empty:
        raise RuntimeError(f"No successful run rows found under {training_root}")
    out["benchmark_order"] = out["benchmark"].map(_benchmark_sort_key)
    out["profile_order"] = out["profile"].map(_profile_sort_key)
    out = out.sort_values(
        ["benchmark_order", "benchmark", "family_label", "profile_order", "repeat_idx", "run_name"],
        kind="stable",
    ).reset_index(drop=True)
    return out


def build_benchmark_profile_summary(all_runs_df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        all_runs_df.groupby(
            ["benchmark", "family", "family_label", "profile"],
            dropna=False,
        )
        .agg(
            n_runs=("run_name", "size"),
            mean_test_f1=("test_f1", "mean"),
            std_test_f1=("test_f1", "std"),
            mean_test_precision=("test_precision", "mean"),
            std_test_precision=("test_precision", "std"),
            mean_test_recall=("test_recall", "mean"),
            std_test_recall=("test_recall", "std"),
            mean_test_accuracy=("test_accuracy", "mean"),
            std_test_accuracy=("test_accuracy", "std"),
            mean_best_val_f1=("best_val_f1", "mean"),
            std_best_val_f1=("best_val_f1", "std"),
            mean_train_rows=("train_rows", "mean"),
        )
        .reset_index()
    )
    summary["std_test_f1"] = summary["std_test_f1"].fillna(0.0)
    summary["std_test_precision"] = summary["std_test_precision"].fillna(0.0)
    summary["std_test_recall"] = summary["std_test_recall"].fillna(0.0)
    summary["std_test_accuracy"] = summary["std_test_accuracy"].fillna(0.0)
    summary["std_best_val_f1"] = summary["std_best_val_f1"].fillna(0.0)
    summary["benchmark_order"] = summary["benchmark"].map(_benchmark_sort_key)
    summary["profile_order"] = summary["profile"].map(_profile_sort_key)
    baseline = (
        summary[(summary["family"] == "ditto_baseline") & (summary["profile"] == "official")][
            ["benchmark", "mean_test_f1"]
        ]
        .rename(columns={"mean_test_f1": "baseline_mean_test_f1"})
        .copy()
    )
    summary = summary.merge(baseline, on="benchmark", how="left")
    summary["delta_mean_f1_vs_baseline"] = summary["mean_test_f1"] - summary["baseline_mean_test_f1"]
    summary["f1_mean_std"] = summary.apply(
        lambda row: f"{row['mean_test_f1']:.4f} ± {row['std_test_f1']:.4f}",
        axis=1,
    )
    return summary.sort_values(
        ["benchmark_order", "benchmark", "family_label", "profile_order"],
        kind="stable",
    ).drop(columns=["benchmark_order", "profile_order"]).reset_index(drop=True)


def build_method_summary(benchmark_profile_df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        benchmark_profile_df.groupby(["family", "family_label", "profile"], dropna=False)
        .agg(
            benchmarks_tested=("benchmark", "nunique"),
            benchmark_list=("benchmark", lambda s: ", ".join(sorted(set(str(v) for v in s)))),
            avg_mean_test_f1=("mean_test_f1", "mean"),
            avg_std_test_f1=("std_test_f1", "mean"),
            best_mean_test_f1=("mean_test_f1", "max"),
            worst_mean_test_f1=("mean_test_f1", "min"),
            avg_delta_mean_f1_vs_baseline=("delta_mean_f1_vs_baseline", "mean"),
            total_runs=("n_runs", "sum"),
        )
        .reset_index()
    )
    summary["f1_mean_std"] = summary.apply(
        lambda row: f"{row['avg_mean_test_f1']:.4f} ± {row['avg_std_test_f1']:.4f}",
        axis=1,
    )
    summary["profile_order"] = summary["profile"].map(_profile_sort_key)
    return summary.sort_values(
        ["avg_mean_test_f1", "benchmarks_tested", "profile_order"],
        ascending=[False, False, True],
        kind="stable",
    ).drop(columns=["profile_order"]).reset_index(drop=True)


def build_method_matrix(benchmark_profile_df: pd.DataFrame) -> pd.DataFrame:
    matrix_source = benchmark_profile_df.copy()
    matrix = matrix_source.pivot_table(
        index=["family_label", "profile"],
        columns="benchmark",
        values="f1_mean_std",
        aggfunc="first",
    ).reset_index()
    ordered_cols = ["family_label", "profile"] + [
        benchmark for benchmark, _ in sorted(BENCHMARK_ORDER.items(), key=lambda item: item[1]) if benchmark in matrix.columns
    ]
    return matrix[ordered_cols]


def build_selected_runs_summary(all_runs_df: pd.DataFrame, benchmark_profile_df: pd.DataFrame) -> pd.DataFrame:
    selected_profiles = benchmark_profile_df.copy()
    selected_profiles["benchmark_order"] = selected_profiles["benchmark"].map(_benchmark_sort_key)
    selected_profiles["profile_order"] = selected_profiles["profile"].map(_profile_sort_key)
    selected_profiles = selected_profiles.sort_values(
        ["benchmark_order", "benchmark", "family", "mean_test_f1", "profile_order"],
        ascending=[True, True, True, False, True],
        kind="stable",
    )
    selected_profiles = selected_profiles.drop_duplicates(
        subset=["family", "benchmark"],
        keep="first",
    ).copy()

    candidate_runs = all_runs_df.merge(
        selected_profiles[
            [
                "family",
                "benchmark",
                "profile",
                "mean_test_f1",
                "std_test_f1",
                "n_runs",
            ]
        ],
        on=["family", "benchmark", "profile"],
        how="inner",
    )
    candidate_runs = candidate_runs.sort_values(
        ["benchmark_order", "benchmark", "family", "test_f1", "repeat_idx", "run_name"],
        ascending=[True, True, True, False, True, True],
        kind="stable",
    )
    selected_runs = candidate_runs.drop_duplicates(subset=["family", "benchmark"], keep="first").copy()
    baseline = (
        benchmark_profile_df[
            (benchmark_profile_df["family"] == "ditto_baseline")
            & (benchmark_profile_df["profile"] == "official")
        ][["benchmark", "mean_test_f1"]]
        .rename(columns={"mean_test_f1": "baseline_avg_f1_over_3_runs"})
        .copy()
    )
    selected_runs = selected_runs.merge(baseline, on="benchmark", how="left")
    selected_runs["delta_vs_baseline_avg_f1"] = (
        selected_runs["mean_test_f1"] - selected_runs["baseline_avg_f1_over_3_runs"]
    )
    selected_runs = selected_runs.rename(
        columns={
            "family": "method",
            "profile": "selected_profile",
            "mean_test_f1": "avg_f1_over_3_runs",
            "std_test_f1": "std_f1_over_3_runs",
            "run_name": "selected_run",
            "test_f1": "selected_run_test_f1",
        }
    )
    selected_runs["method_label"] = selected_runs["method"].map(_family_label)
    ordered_columns = [
        "method",
        "method_label",
        "benchmark",
        "selected_profile",
        "avg_f1_over_3_runs",
        "std_f1_over_3_runs",
        "baseline_avg_f1_over_3_runs",
        "delta_vs_baseline_avg_f1",
        "n_runs",
        "selected_run",
        "selected_run_test_f1",
        "repeat_idx",
        "seed",
    ]
    return selected_runs[ordered_columns].sort_values(
        ["benchmark", "method_label"],
        key=lambda col: col.map(_benchmark_sort_key) if col.name == "benchmark" else col,
        kind="stable",
    ).reset_index(drop=True)


def _apply_workbook_formatting(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    positive_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    negative_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    percent_cols = {
        "test_f1",
        "test_precision",
        "test_recall",
        "test_accuracy",
        "best_val_f1",
        "mean_test_f1",
        "std_test_f1",
        "mean_test_precision",
        "std_test_precision",
        "mean_test_recall",
        "std_test_recall",
        "mean_test_accuracy",
        "std_test_accuracy",
        "mean_best_val_f1",
        "std_best_val_f1",
        "baseline_mean_test_f1",
        "baseline_avg_f1_over_3_runs",
        "avg_mean_test_f1",
        "avg_std_test_f1",
        "best_mean_test_f1",
        "worst_mean_test_f1",
        "avg_f1_over_3_runs",
        "std_f1_over_3_runs",
        "selected_run_test_f1",
    }
    delta_cols = {"delta_mean_f1_vs_baseline", "avg_delta_mean_f1_vs_baseline", "delta_vs_baseline_avg_f1"}

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        headers = [cell.value for cell in sheet[1]]
        header_map = {idx + 1: str(value) for idx, value in enumerate(headers) if value is not None}
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for col_idx, header in header_map.items():
            max_len = len(header)
            for cell in sheet[get_column_letter(col_idx)]:
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
                if cell.row == 1:
                    continue
                if header in percent_cols and isinstance(value, (int, float)):
                    cell.number_format = "0.0%"
                elif header in delta_cols and isinstance(value, (int, float)):
                    cell.number_format = "+0.0%;-0.0%;0.0%"
                    if value > 0:
                        cell.fill = positive_fill
                    elif value < 0:
                        cell.fill = negative_fill
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)

    workbook.save(path)


def write_workbook(
    output_path: Path,
    selected_runs_df: pd.DataFrame,
    method_summary_df: pd.DataFrame,
    method_matrix_df: pd.DataFrame,
    benchmark_profile_df: pd.DataFrame,
    all_runs_df: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        selected_runs_df.to_excel(writer, sheet_name="Selected Runs", index=False)
        method_summary_df.to_excel(writer, sheet_name="Method Summary", index=False)
        method_matrix_df.to_excel(writer, sheet_name="Method Matrix", index=False)
        benchmark_profile_df.to_excel(writer, sheet_name="Benchmark Profile", index=False)
        all_runs_df.drop(columns=["benchmark_order", "profile_order"]).to_excel(
            writer,
            sheet_name="All Runs",
            index=False,
        )
    _apply_workbook_formatting(output_path)


def main() -> None:
    args = _parse_args()
    export_dir = Path(args.export_dir).expanduser().resolve()
    if not export_dir.exists():
        raise FileNotFoundError(f"Export directory does not exist: {export_dir}")

    output_path = Path(args.output).expanduser().resolve() if args.output else export_dir / DEFAULT_OUTPUT_NAME
    all_runs_df = collect_runs(export_dir)
    benchmark_profile_df = build_benchmark_profile_summary(all_runs_df)
    selected_runs_df = build_selected_runs_summary(all_runs_df, benchmark_profile_df)
    method_summary_df = build_method_summary(benchmark_profile_df)
    method_matrix_df = build_method_matrix(benchmark_profile_df)
    write_workbook(
        output_path=output_path,
        selected_runs_df=selected_runs_df,
        method_summary_df=method_summary_df,
        method_matrix_df=method_matrix_df,
        benchmark_profile_df=benchmark_profile_df,
        all_runs_df=all_runs_df,
    )
    print(output_path)


if __name__ == "__main__":
    main()
