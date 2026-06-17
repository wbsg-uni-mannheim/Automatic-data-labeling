#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[3]
DEFAULT_EXPORT_DIR = ROOT / "export_261304"
DEFAULT_OUTPUT_NAME = "results_method_overview_compact.xlsx"

BENCHMARK_ORDER = {
    "abt-buy": 0,
    "amazon-google": 1,
    "dblp-acm": 2,
    "dblp-scholar": 3,
    "walmart-amazon": 4,
    "wdc": 5,
}
METHODS = [
    "Baseline",
    "Traditional Random Search",
    "Three-Stage Active Learning",
    "Three-Stage Active Learning + Re-Labeling",
]
PROFILE_ORDER = [
    "official",
    "small",
    "small_plus20random",
    "medium",
    "medium_plus20random",
    "large",
    "all",
    "all_plus20random",
]


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build a one-sheet Excel overview with one best result per benchmark and "
            "method, including the labeling model used."
        )
    )
    parser.add_argument(
        "export_dir",
        nargs="?",
        default=str(DEFAULT_EXPORT_DIR),
        help="Path to the exported results directory. Default: %(default)s",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Target xlsx path. Default: <export_dir>/results_method_overview_compact.xlsx",
    )
    return parser.parse_args()


def _safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text())


def _profile_rank(profile: Any) -> int:
    profile_str = str(profile)
    try:
        return PROFILE_ORDER.index(profile_str)
    except ValueError:
        return 999


def _extract_profile_from_run_name(run_name: str) -> str:
    for candidate in sorted(PROFILE_ORDER, key=len, reverse=True):
        suffix = f"_{candidate}_"
        if suffix in run_name:
            return candidate
    return "unknown"


def _collect_manifest_by_benchmark(root: Path) -> Dict[str, Dict[str, Any]]:
    manifests: Dict[str, Dict[str, Any]] = {}
    if not root.exists():
        return manifests
    for manifest_path in sorted(root.glob("*/profile_manifest.json")):
        manifest = _load_json(manifest_path)
        benchmark = str(manifest.get("benchmark") or "")
        if benchmark:
            manifests[benchmark] = manifest
    return manifests


def _default_model_for_family(manifests: Dict[str, Dict[str, Any]]) -> Optional[str]:
    for manifest in manifests.values():
        model = ((manifest.get("labeling_cost") or {}).get("model")) or None
        if model:
            return str(model)
    return None


def _format_relabel_model(
    relabel_manifest: Dict[str, Any],
    source_manifest: Optional[Dict[str, Any]],
    default_relabel_model: Optional[str] = None,
    default_source_model: Optional[str] = None,
) -> str:
    source_model = ((source_manifest or {}).get("labeling_cost") or {}).get("model") or default_source_model
    relabel_model = (relabel_manifest.get("labeling_cost") or {}).get("model") or default_relabel_model
    if source_model and relabel_model and source_model != relabel_model:
        return f"{source_model} -> {relabel_model}"
    if relabel_model:
        return str(relabel_model)
    if source_model:
        return str(source_model)
    return "n/a"


def collect_family_models(
    export_dir: Path,
    family: str,
    source_family: Optional[str] = None,
) -> Dict[str, str]:
    family_manifests = _collect_manifest_by_benchmark(export_dir / family)
    source_manifests = (
        _collect_manifest_by_benchmark(export_dir / source_family) if source_family else {}
    )
    default_family_model = _default_model_for_family(family_manifests)
    default_source_model = _default_model_for_family(source_manifests)
    models: Dict[str, str] = {}

    for benchmark, manifest in family_manifests.items():
        if source_family:
            models[benchmark] = _format_relabel_model(
                manifest,
                source_manifests.get(benchmark),
                default_relabel_model=default_family_model,
                default_source_model=default_source_model,
            )
        else:
            models[benchmark] = str(
                ((manifest.get("labeling_cost") or {}).get("model")) or default_family_model or "n/a"
            )
    return models


def collect_baselines(export_dir: Path) -> Dict[str, Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for summary_path in sorted((export_dir / "ditto_benchmark_runs").glob("*/summary.csv")):
        df = pd.read_csv(summary_path)
        for row in df.to_dict("records"):
            test_f1 = _safe_float(row.get("test_f1"))
            if test_f1 is None:
                continue
            rows.append(
                {
                    "benchmark": str(row["benchmark"]),
                    "labeling_model": "gold labels",
                    "profile": "official",
                    "f1": test_f1,
                    "source": str(summary_path),
                    "note": "Official benchmark run on gold labels.",
                }
            )

    frame = pd.DataFrame(rows)
    if frame.empty:
        raise RuntimeError("No baseline F1 scores found.")
    frame = frame.sort_values(["benchmark", "f1", "source"], ascending=[True, False, False])
    frame = frame.drop_duplicates("benchmark", keep="first")
    return {row["benchmark"]: row for row in frame.to_dict("records")}


def collect_traditional_random_search(export_dir: Path) -> Dict[str, Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    root = export_dir / "training_from_generated_labels"
    if not root.exists():
        return {}

    for family_dir in sorted(path for path in root.iterdir() if path.is_dir()):
        family_name = family_dir.name.lower()
        if "random" not in family_name and "level" not in family_name:
            continue
        for summary_path in sorted(family_dir.rglob("summary.csv")):
            df = pd.read_csv(summary_path)
            if df.empty:
                continue
            row = df.iloc[0].to_dict()
            if str(row.get("status", "")).lower() != "ok":
                continue

            rows.append(
                {
                    "benchmark": str(row.get("benchmark", "")),
                    "labeling_model": "unknown",
                    "profile": _extract_profile_from_run_name(summary_path.parent.name),
                    "f1": _safe_float(row.get("test_f1")),
                    "source": str(summary_path),
                    "note": "Candidate random-only training run found in export.",
                }
            )

    frame = pd.DataFrame(rows)
    if frame.empty:
        return {}

    frame = frame.dropna(subset=["f1"]).copy()
    frame["profile_rank"] = frame["profile"].map(_profile_rank)
    frame = frame.sort_values(
        ["benchmark", "f1", "profile_rank", "source"],
        ascending=[True, False, True, False],
        kind="stable",
    ).drop(columns=["profile_rank"])
    frame = frame.drop_duplicates("benchmark", keep="first")
    return {row["benchmark"]: row for row in frame.to_dict("records")}


def collect_family_best(
    export_dir: Path,
    family: str,
    method: str,
    note: str,
    source_family: Optional[str] = None,
) -> Dict[str, Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    root = export_dir / "training_from_generated_labels" / family
    if not root.exists():
        return {}

    family_manifests = _collect_manifest_by_benchmark(export_dir / family)
    source_manifests = (
        _collect_manifest_by_benchmark(export_dir / source_family) if source_family else {}
    )
    default_family_model = _default_model_for_family(family_manifests)
    default_source_model = _default_model_for_family(source_manifests)

    for summary_path in sorted(root.rglob("summary.csv")):
        df = pd.read_csv(summary_path)
        if df.empty:
            continue
        row = df.iloc[0].to_dict()
        if str(row.get("status", "")).lower() != "ok":
            continue

        benchmark = str(row.get("benchmark", ""))
        if not benchmark:
            continue

        manifest = family_manifests.get(benchmark, {})
        if source_family:
            labeling_model = _format_relabel_model(
                manifest,
                source_manifests.get(benchmark),
                default_relabel_model=default_family_model,
                default_source_model=default_source_model,
            )
        else:
            labeling_model = str(((manifest.get("labeling_cost") or {}).get("model")) or default_family_model or "n/a")

        rows.append(
            {
                "benchmark": benchmark,
                "labeling_model": labeling_model,
                "profile": _extract_profile_from_run_name(summary_path.parent.name),
                "f1": _safe_float(row.get("test_f1")),
                "source": str(summary_path),
                "note": note,
                "method": method,
            }
        )

    frame = pd.DataFrame(rows)
    if frame.empty:
        return {}
    frame = frame.dropna(subset=["f1"]).copy()
    frame["profile_rank"] = frame["profile"].map(_profile_rank)
    frame = frame.sort_values(
        ["benchmark", "f1", "profile_rank", "source"],
        ascending=[True, False, True, False],
        kind="stable",
    ).drop(columns=["profile_rank"])
    frame = frame.drop_duplicates("benchmark", keep="first")
    return {row["benchmark"]: row for row in frame.to_dict("records")}


def build_compact_rows(export_dir: Path) -> pd.DataFrame:
    baselines = collect_baselines(export_dir)
    traditional_random = collect_traditional_random_search(export_dir)
    three_stage_models = collect_family_models(
        export_dir,
        family="three_phase_labeling_ditto_only_v2",
    )
    three_stage_relabel_models = collect_family_models(
        export_dir,
        family="three_phase_labeling_ditto_only_v2_relabel_batch_gpt-5-mini_agent_precision",
        source_family="three_phase_labeling_ditto_only_v2",
    )
    three_stage = collect_family_best(
        export_dir,
        family="three_phase_labeling_ditto_only_v2",
        method="Three-Stage Active Learning",
        note="Best downstream Ditto run from the three-stage labeling family.",
    )
    three_stage_relabel = collect_family_best(
        export_dir,
        family="three_phase_labeling_ditto_only_v2_relabel_batch_gpt-5-mini_agent_precision",
        method="Three-Stage Active Learning + Re-Labeling",
        note="Best downstream Ditto run after batch re-labeling.",
        source_family="three_phase_labeling_ditto_only_v2",
    )

    datasets = sorted(baselines, key=lambda key: BENCHMARK_ORDER.get(key, 999))
    rows: List[Dict[str, Any]] = []

    method_maps = {
        "Baseline": baselines,
        "Traditional Random Search": traditional_random,
        "Three-Stage Active Learning": three_stage,
        "Three-Stage Active Learning + Re-Labeling": three_stage_relabel,
    }

    for dataset in datasets:
        baseline_f1 = _safe_float(baselines[dataset]["f1"])
        for method in METHODS:
            entry = method_maps[method].get(dataset)

            if entry is None:
                if method == "Traditional Random Search":
                    note = (
                        "Im Export wurde kein separater Random-Only-Labelinglauf mit "
                        "nachgelagertem F1 gefunden."
                    )
                    labeling_model = "n/a"
                elif method == "Three-Stage Active Learning":
                    note = (
                        "Im Export liegt fuer diesen Datensatz kein downstream Ditto-F1 "
                        "unter training_from_generated_labels/three_phase_labeling_ditto_only_v2 vor."
                    )
                    labeling_model = three_stage_models.get(dataset, "n/a")
                elif method == "Three-Stage Active Learning + Re-Labeling":
                    note = (
                        "Im Export liegt fuer diesen Datensatz kein erfolgreicher downstream "
                        "Ditto-F1 unter training_from_generated_labels/"
                        "three_phase_labeling_ditto_only_v2_relabel_batch_gpt-5-mini_agent_precision vor."
                    )
                    labeling_model = three_stage_relabel_models.get(dataset, "n/a")
                else:
                    note = "Kein Ergebnis im Export vorhanden."
                    labeling_model = "gold labels"

                rows.append(
                    {
                        "dataset": dataset,
                        "method": method,
                        "labeling_model": labeling_model,
                        "best_profile": None,
                        "f1": None,
                        "delta_vs_baseline": None,
                        "note": note,
                    }
                )
                continue

            f1 = _safe_float(entry.get("f1"))
            rows.append(
                {
                    "dataset": dataset,
                    "method": method,
                    "labeling_model": entry.get("labeling_model"),
                    "best_profile": entry.get("profile"),
                    "f1": f1,
                    "delta_vs_baseline": (f1 - baseline_f1) if f1 is not None else None,
                    "note": entry.get("note"),
                }
            )

        rows.append(
            {
                "dataset": None,
                "method": None,
                "labeling_model": None,
                "best_profile": None,
                "f1": None,
                "delta_vs_baseline": None,
                "note": None,
            }
        )

    return pd.DataFrame(rows)


def write_workbook(output_path: Path, df: pd.DataFrame) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Overview", index=False)

    workbook = load_workbook(output_path)
    sheet = workbook["Overview"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    positive_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    negative_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    dataset_fill = PatternFill(fill_type="solid", fgColor="F3F3F3")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in sheet.iter_rows(min_row=2):
        dataset_cell = row[0]
        method_cell = row[1]
        f1_cell = row[4]
        delta_cell = row[5]
        if dataset_cell.value and method_cell.value == "Baseline":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = dataset_fill
        if isinstance(f1_cell.value, (int, float)):
            f1_cell.number_format = "0.0%"
        if isinstance(delta_cell.value, (int, float)):
            delta_cell.number_format = "+0.0%;-0.0%;0.0%"
            if delta_cell.value > 0:
                delta_cell.fill = positive_fill
            elif delta_cell.value < 0:
                delta_cell.fill = negative_fill

    for idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = 12
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 70)

    workbook.save(output_path)


def main() -> None:
    args = _parse_args()
    export_dir = Path(args.export_dir).expanduser().resolve()
    if not export_dir.exists():
        raise FileNotFoundError(f"Export directory does not exist: {export_dir}")

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else export_dir / DEFAULT_OUTPUT_NAME
    )

    df = build_compact_rows(export_dir)
    write_workbook(output_path, df)
    print(output_path)


if __name__ == "__main__":
    main()
