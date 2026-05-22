#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXPORT_DIR = ROOT / "export_260414_0938"
DEFAULT_BASELINE_EXPORT_DIR = ROOT / "export_261304"
DEFAULT_ACTIVE_LEARNING_DIR = ROOT / "output" / "autolabel_v1"
DEFAULT_OUTPUT_NAME = "results_best_by_method.xlsx"

BENCHMARK_ORDER = {
    "abt-buy": 0,
    "amazon-google": 1,
    "dblp-acm": 2,
    "dblp-scholar": 3,
    "walmart-amazon": 4,
    "wdc": 5,
}
PROFILE_ORDER = {
    "official": 0,
    "small": 1,
    "small_plus20random": 2,
    "medium": 3,
    "medium_plus20random": 4,
    "large": 5,
    "all": 6,
    "all_plus20random": 7,
    "legacy": 8,
}
METHOD_ORDER = {
    "Baseline": 0,
    "Active Learning v1": 1,
    "Seed Round Only": 2,
    "Simple Active Learning": 3,
    "Three-Phase Active Learning v2": 4,
    "Three-Phase Ditto-only v2": 5,
    "Three-Phase v2 + Batch Relabel": 6,
    "Three-Phase v2 + Drop Changed": 7,
    "Three-Phase v2 + Closure Bridge Drop": 8,
    "Three-Phase v2 + Closure Bridge + Relabel-Changed Drop": 9,
    "Three-Phase v2 + Closure Bridge OR Relabel-Changed Drop": 10,
}
FAMILY_TO_METHOD = {
    "active_learning_v1": "Active Learning v1",
    "seed_round_only_profiles": "Seed Round Only",
    "simple_active_learning_labeling": "Simple Active Learning",
    "three_phase_active_learning_v2": "Three-Phase Active Learning v2",
    "three_phase_labeling_ditto_only_v2": "Three-Phase Ditto-only v2",
    "three_phase_labeling_ditto_only_v2_drop_changed": "Three-Phase v2 + Drop Changed",
    "three_phase_labeling_ditto_only_v2_closure_bridge_drop": (
        "Three-Phase v2 + Closure Bridge Drop"
    ),
    "three_phase_labeling_ditto_only_v2_closure_bridge_relabel_changed_drop": (
        "Three-Phase v2 + Closure Bridge + Relabel-Changed Drop"
    ),
    "three_phase_labeling_ditto_only_v2_closure_bridge_or_relabel_changed_drop": (
        "Three-Phase v2 + Closure Bridge OR Relabel-Changed Drop"
    ),
    "three_phase_labeling_ditto_only_v2_relabel_batch_gpt-5-mini_agent_precision": (
        "Three-Phase v2 + Batch Relabel"
    ),
}


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a compact Excel workbook with one tab per dataset and only the best result per method."
    )
    parser.add_argument(
        "export_dir",
        nargs="?",
        default=str(DEFAULT_EXPORT_DIR),
        help="Path to the export directory with training results. Default: %(default)s",
    )
    parser.add_argument(
        "--baseline-export-dir",
        default=str(DEFAULT_BASELINE_EXPORT_DIR),
        help="Export directory that contains ditto_benchmark_runs. Default: %(default)s",
    )
    parser.add_argument(
        "--active-learning-dir",
        default=str(DEFAULT_ACTIVE_LEARNING_DIR),
        help="Directory with legacy active learning Ditto results. Default: %(default)s",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Target .xlsx path. Default: <export_dir>/results_best_by_method.xlsx",
    )
    return parser.parse_args()


def _safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text())


def _parse_profile_from_run_name(run_name: str) -> str:
    for profile in sorted(PROFILE_ORDER, key=len, reverse=True):
        if re.search(rf"_{re.escape(profile)}_\d{{8}}_\d{{6}}$", run_name):
            return profile
    return "unknown"


def _profile_rank(profile: Any) -> int:
    return PROFILE_ORDER.get(str(profile), 99)


def _collect_baselines(baseline_export_dir: Path) -> Dict[str, Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    candidate_paths: List[Path] = []
    ditto_benchmark_runs = baseline_export_dir / "ditto_benchmark_runs"
    if ditto_benchmark_runs.exists():
        candidate_paths.extend(sorted(ditto_benchmark_runs.glob("*/summary.csv")))
    ditto_baseline_summary = baseline_export_dir / "ditto_baseline" / "summary.csv"
    if ditto_baseline_summary.exists():
        candidate_paths.append(ditto_baseline_summary)

    for summary_path in candidate_paths:
        df = pd.read_csv(summary_path)
        for row in df.to_dict("records"):
            if str(row.get("status", "ok")).lower() != "ok":
                continue
            test_f1 = _safe_float(row.get("test_f1"))
            if test_f1 is None:
                continue
            rows.append(
                {
                    "benchmark": str(row.get("benchmark")),
                    "method": "Baseline",
                    "best_profile": "official",
                    "f1": test_f1,
                    "delta_vs_baseline": 0.0,
                    "source": str(summary_path),
                }
            )
    frame = pd.DataFrame(rows)
    if frame.empty:
        raise RuntimeError(
            f"No baseline F1 values found in {baseline_export_dir / 'ditto_benchmark_runs'} "
            f"or {baseline_export_dir / 'ditto_baseline' / 'summary.csv'}"
        )
    frame = frame.sort_values(["benchmark", "f1", "source"], ascending=[True, False, False], kind="stable")
    frame = frame.drop_duplicates("benchmark", keep="first")
    return {row["benchmark"]: row for row in frame.to_dict("records")}


def _collect_export_family_best(
    export_dir: Path,
    baseline_index: Dict[str, Dict[str, Any]],
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    family_roots: List[Path] = []
    nested_root = export_dir / "training_from_generated_labels"
    if nested_root.exists():
        family_roots.extend(sorted(path for path in nested_root.iterdir() if path.is_dir()))
    family_roots.extend(sorted(path for path in export_dir.iterdir() if path.is_dir()))

    seen_family_dirs: set[Path] = set()
    for family_dir in family_roots:
        if family_dir in seen_family_dirs:
            continue
        seen_family_dirs.add(family_dir)
        family = family_dir.name
        method = FAMILY_TO_METHOD.get(family)
        if not method:
            continue
        for summary_path in sorted(family_dir.rglob("summary.csv")):
            df = pd.read_csv(summary_path)
            if df.empty:
                continue
            row = df.iloc[0].to_dict()
            if str(row.get("status", "")).lower() != "ok":
                continue
            benchmark = str(row.get("benchmark", "")).strip()
            baseline = baseline_index.get(benchmark, {})
            f1 = _safe_float(row.get("test_f1"))
            if not benchmark or f1 is None:
                continue
            rows.append(
                {
                    "benchmark": benchmark,
                    "method": method,
                    "best_profile": _parse_profile_from_run_name(summary_path.parent.name),
                    "f1": f1,
                    "delta_vs_baseline": (
                        f1 - _safe_float(baseline.get("f1"))
                        if _safe_float(baseline.get("f1")) is not None
                        else None
                    ),
                    "source": str(summary_path),
                }
            )
    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    frame["profile_rank"] = frame["best_profile"].map(_profile_rank)
    frame = frame.sort_values(
        ["benchmark", "method", "f1", "profile_rank", "source"],
        ascending=[True, True, False, True, False],
        kind="stable",
    ).drop(columns=["profile_rank"])
    frame = frame.drop_duplicates(["benchmark", "method"], keep="first")
    return frame.reset_index(drop=True)


def _collect_active_learning_v1(
    active_learning_dir: Path,
    baseline_index: Dict[str, Dict[str, Any]],
) -> pd.DataFrame:
    if not active_learning_dir.exists():
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    method = "Active Learning v1"

    for metrics_path in sorted(active_learning_dir.glob("*/metrics.tsv")):
        df = pd.read_csv(metrics_path, sep="\t")
        for row in df.to_dict("records"):
            benchmark = str(row.get("benchmark"))
            f1 = _safe_float(row.get("test_f1"))
            baseline = baseline_index.get(benchmark, {})
            rows.append(
                {
                    "benchmark": benchmark,
                    "method": method,
                    "best_profile": str(row.get("profile")),
                    "f1": f1,
                    "delta_vs_baseline": (
                        f1 - _safe_float(baseline.get("f1"))
                        if f1 is not None and _safe_float(baseline.get("f1")) is not None
                        else None
                    ),
                    "source": str(metrics_path),
                }
            )

    abt_root = active_learning_dir / "abt_buy_profiles_abt_buy_local_test_173424"
    for metrics_json in sorted(abt_root.rglob("metrics.json")):
        profile = metrics_json.parents[1].name
        payload = _load_json(metrics_json)
        test = payload.get("test") or {}
        benchmark = "abt-buy"
        f1 = _safe_float(test.get("f1"))
        baseline = baseline_index.get(benchmark, {})
        rows.append(
            {
                "benchmark": benchmark,
                "method": method,
                "best_profile": profile,
                "f1": f1,
                "delta_vs_baseline": (
                    f1 - _safe_float(baseline.get("f1"))
                    if f1 is not None and _safe_float(baseline.get("f1")) is not None
                    else None
                ),
                "source": str(metrics_json),
            }
        )

    wdc_metrics = active_learning_dir / "wdc" / "run_20260220_235713" / "metrics.json"
    if wdc_metrics.exists():
        payload = _load_json(wdc_metrics)
        test = payload.get("test") or {}
        benchmark = "wdc"
        f1 = _safe_float(test.get("f1"))
        baseline = baseline_index.get(benchmark, {})
        rows.append(
            {
                "benchmark": benchmark,
                "method": method,
                "best_profile": "legacy",
                "f1": f1,
                "delta_vs_baseline": (
                    f1 - _safe_float(baseline.get("f1"))
                    if f1 is not None and _safe_float(baseline.get("f1")) is not None
                    else None
                ),
                "source": str(wdc_metrics),
            }
        )

    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    frame["profile_rank"] = frame["best_profile"].map(_profile_rank)
    frame = frame.sort_values(
        ["benchmark", "method", "f1", "profile_rank", "source"],
        ascending=[True, True, False, True, False],
        kind="stable",
    ).drop(columns=["profile_rank"])
    frame = frame.drop_duplicates(["benchmark", "method"], keep="first")
    return frame.reset_index(drop=True)


def _build_dataset_frames(
    baseline_index: Dict[str, Dict[str, Any]],
    export_best_df: pd.DataFrame,
    active_learning_df: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    method_frames = []
    if not export_best_df.empty:
        method_frames.append(export_best_df)
    if not active_learning_df.empty:
        method_frames.append(active_learning_df)

    all_methods_df = pd.concat(method_frames, ignore_index=True, sort=False) if method_frames else pd.DataFrame()
    dataset_frames: Dict[str, pd.DataFrame] = {}
    available_methods = ["Baseline"]
    if not all_methods_df.empty:
        present_methods = [
            method
            for method in METHOD_ORDER
            if method != "Baseline" and method in set(all_methods_df["method"].dropna().astype(str))
        ]
        available_methods.extend(present_methods)

    for benchmark in sorted(baseline_index, key=lambda x: BENCHMARK_ORDER.get(x, 99)):
        rows: List[Dict[str, Any]] = [baseline_index[benchmark]]
        by_method: Dict[str, Dict[str, Any]] = {}
        if not all_methods_df.empty:
            subset = all_methods_df[all_methods_df["benchmark"] == benchmark].copy()
            if not subset.empty:
                subset["method_order"] = subset["method"].map(lambda x: METHOD_ORDER.get(str(x), 99))
                subset = subset.sort_values(["method_order", "f1"], ascending=[True, False], kind="stable")
                subset = subset.drop(columns=["method_order"])
                by_method = {row["method"]: row for row in subset.to_dict("records")}
        for method in [m for m in available_methods if m != "Baseline"]:
            if method in by_method:
                rows.append(by_method[method])
        frame = pd.DataFrame(rows)
        if not frame.empty:
            frame = frame[["method", "best_profile", "f1", "delta_vs_baseline", "source"]]
        dataset_frames[benchmark] = frame
    return dataset_frames


def _write_workbook(path: Path, dataset_frames: Dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for benchmark, frame in dataset_frames.items():
            frame.to_excel(writer, sheet_name=benchmark[:31], index=False)

    workbook = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    positive_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    negative_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    baseline_fill = PatternFill(fill_type="solid", fgColor="F3F3F3")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        headers = [cell.value for cell in sheet[1]]
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row in sheet.iter_rows(min_row=2):
            method_cell = row[0]
            f1_cell = row[2]
            delta_cell = row[3]
            if method_cell.value == "Baseline":
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = baseline_fill
            if isinstance(f1_cell.value, (int, float)):
                f1_cell.number_format = "0.0%"
            if isinstance(delta_cell.value, (int, float)):
                delta_cell.number_format = "+0.0%;-0.0%;0.0%"
                if delta_cell.value > 0:
                    delta_cell.fill = positive_fill
                elif delta_cell.value < 0:
                    delta_cell.fill = negative_fill

        for idx, col in enumerate(sheet.columns, start=1):
            max_len = 12
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    workbook.save(path)


def main() -> None:
    args = _parse_args()
    export_dir = Path(args.export_dir).expanduser().resolve()
    baseline_export_dir = Path(args.baseline_export_dir).expanduser().resolve()
    active_learning_dir = Path(args.active_learning_dir).expanduser().resolve()

    if not export_dir.exists():
        raise FileNotFoundError(f"Export directory does not exist: {export_dir}")
    if not baseline_export_dir.exists():
        raise FileNotFoundError(f"Baseline export directory does not exist: {baseline_export_dir}")

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else export_dir / DEFAULT_OUTPUT_NAME
    )

    baseline_index = _collect_baselines(baseline_export_dir)
    export_best_df = _collect_export_family_best(export_dir, baseline_index)
    active_learning_df = _collect_active_learning_v1(active_learning_dir, baseline_index)
    dataset_frames = _build_dataset_frames(baseline_index, export_best_df, active_learning_df)
    _write_workbook(output_path, dataset_frames)
    print(output_path)


if __name__ == "__main__":
    main()
