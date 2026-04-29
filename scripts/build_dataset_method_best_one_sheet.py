#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "scripts") not in sys.path:
    sys.path.insert(0, str(ROOT / "scripts"))

from build_dataset_method_best_workbook import (
    BENCHMARK_ORDER,
    DEFAULT_ACTIVE_LEARNING_DIR,
    DEFAULT_BASELINE_EXPORT_DIR,
    DEFAULT_EXPORT_DIR,
    _build_dataset_frames,
    _collect_active_learning_v1,
    _collect_baselines,
    _collect_export_family_best,
)
DEFAULT_OUTPUT_NAME = "results_best_by_method_one_sheet.xlsx"
SIMPLE_LABELING_SYSTEM_PROMPT = (
    "You are an expert entity matcher. Decide if two records refer to the same real-world "
    'entity. Return only valid JSON with exactly one field: {"match": true|false}.'
)
AGENT_PRECISION_PROMPT = (
    "You are an expert entity matcher. Be conservative: predict match=true only when the "
    "evidence strongly supports the same entity and there is no meaningful contradiction. "
    'Return only valid JSON with exactly two fields: {"match": true|false, "confidence": 50-100}.'
)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a one-sheet Excel overview with the best result per dataset and method."
    )
    parser.add_argument(
        "export_dir",
        nargs="?",
        default=str(DEFAULT_EXPORT_DIR),
        help="Path to the export directory with training results. Default: %(default)s",
    )
    parser.add_argument(
        "--baseline-export-dir",
        default=str(DEFAULT_BASELINE_EXPORT_DIR),
        help="Export directory that contains ditto_benchmark_runs. Default: %(default)s",
    )
    parser.add_argument(
        "--active-learning-dir",
        default=str(DEFAULT_ACTIVE_LEARNING_DIR),
        help="Directory with legacy active learning Ditto results. Default: %(default)s",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Target .xlsx path. Default: <export_dir>/results_best_by_method_one_sheet.xlsx",
    )
    return parser.parse_args()


def _build_one_sheet_rows(dataset_frames: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: List[dict] = []
    for benchmark in sorted(dataset_frames, key=lambda x: BENCHMARK_ORDER.get(x, 99)):
        frame = dataset_frames[benchmark]
        for record in frame.to_dict("records"):
            rows.append(
                {
                    "dataset": benchmark,
                    "method": record.get("method"),
                    "best_profile": record.get("best_profile"),
                    "f1": record.get("f1"),
                    "delta_vs_baseline": record.get("delta_vs_baseline"),
                    "source": record.get("source"),
                }
            )
        rows.append(
            {
                "dataset": None,
                "method": None,
                "best_profile": None,
                "f1": None,
                "delta_vs_baseline": None,
                "source": None,
            }
        )
    return pd.DataFrame(rows)


def _build_method_guide_rows(methods_in_overview: List[str]) -> pd.DataFrame:
    method_rows = {
        "Baseline": {
            "method": "Baseline",
            "labeling_model": "n/a",
            "how_it_works": (
                "Gold-label reference. Ditto is trained on the official benchmark training split; "
                "no generated labels are used."
            ),
            "prompts_used": "No LLM prompt; uses benchmark gold labels only.",
            "prompt_source": "n/a",
            "notes": "Reference line for F1 and delta calculations.",
        },
        "Active Learning v1": {
            "method": "Active Learning v1",
            "labeling_model": "not preserved in export",
            "how_it_works": (
                "Legacy active-learning pipeline that creates labeled profiles and then trains Ditto "
                "for small/medium/large/all profile sizes."
            ),
            "prompts_used": (
                "Exact prompt text is not preserved in output/autolabel_v1. The export only contains "
                "training metrics and a small amount of model config."
            ),
            "prompt_source": "not recoverable from output/autolabel_v1 artifacts",
            "notes": "Method is included for result comparison, but prompt provenance is incomplete.",
        },
        "Seed Round Only": {
            "method": "Seed Round Only",
            "labeling_model": "gpt-5.2 by default",
            "how_it_works": (
                "Uses only the simple-active-learning seed-round strategy. It builds nearest-neighbor "
                "candidates, labels per left-query until the per-query positive/negative seed targets "
                "are met, then materializes the usual profiles from that seed-only master set. It does "
                "not run the active-learning loop and does not use Ditto disagreement sampling."
            ),
            "prompts_used": (
                "Uses the same entity-matcher JSON prompt as Simple Active Learning: "
                f"{SIMPLE_LABELING_SYSTEM_PROMPT}"
            ),
            "prompt_source": (
                "scripts/labeling/run_seed_round_only_profiles.py, reusing seed functions from "
                "scripts/labeling/run_simple_labeling.py"
            ),
            "notes": "Pure seed-selection baseline for separating seed sampling from later active-learning gains.",
        },
        "Simple Active Learning": {
            "method": "Simple Active Learning",
            "labeling_model": "gpt-5.2",
            "how_it_works": (
                "Single-stage active-learning pipeline. It builds a seed set and then expands the "
                "labels iteratively with the same prompt until the target profile size is reached."
            ),
            "prompts_used": (
                "Uses the entity-matcher JSON prompt throughout the labeling loop: "
                f"{SIMPLE_LABELING_SYSTEM_PROMPT}"
            ),
            "prompt_source": "scripts/labeling/run_simple_labeling.py (_label_pair)",
            "notes": "No separate Ditto disagreement phase and no relabel pass.",
        },
        "Three-Phase Active Learning v2": {
            "method": "Three-Phase Active Learning v2",
            "labeling_model": "gpt-5.2",
            "how_it_works": (
                "Phase 1 builds a seed set, Phase 2 expands labels with the same active-learning prompt, "
                "and Phase 3 trains a bagged Ditto ensemble to target uncertain pairs before the final "
                "profiles are built."
            ),
            "prompts_used": (
                "Phase 1 and 2 use the entity-matcher JSON prompt: "
                f"{SIMPLE_LABELING_SYSTEM_PROMPT} "
                "Phase 3 uses Ditto disagreement sampling, so no additional LLM prompt is used there."
            ),
            "prompt_source": (
                "scripts/labeling/run_simple_labeling.py (_label_pair), reused by "
                "scripts/labeling/run_three_phase_labeling.py"
            ),
            "notes": "Profiles may also include *_plus20random variants when random add-ons were built.",
        },
        "Three-Phase v2 + Batch Relabel": {
            "method": "Three-Phase v2 + Batch Relabel",
            "labeling_model": "gpt-5.2 -> gpt-5-mini",
            "how_it_works": (
                "Starts from the same three-phase v2 labels, then re-labels the master profile "
                "(all_plus20random) in batch and rebuilds the downstream profiles from the cleaned labels."
            ),
            "prompts_used": (
                "Initial labeling uses the same Phase 1/2 entity-matcher JSON prompt as three-phase v2. "
                "The relabel step uses the conservative agent_precision prompt: "
                f"{AGENT_PRECISION_PROMPT}"
            ),
            "prompt_source": (
                "Initial prompt: scripts/labeling/run_simple_labeling.py (_label_pair). "
                "Relabel prompt: scripts/experiments/evidence_first_abstain/prompts/"
                "agent_precision_system_prompt.txt via "
                "scripts/labeling/relabel_three_phase_generated_labels_batch.py"
            ),
            "notes": "This is the only method in the compact overview that changes labels after the initial run.",
        },
        "Three-Phase v2 + Drop Changed": {
            "method": "Three-Phase v2 + Drop Changed",
            "labeling_model": "gpt-5.2 -> gpt-5-mini signal, labels unchanged",
            "how_it_works": (
                "Starts from the original three-phase v2 labels and uses the batch-relabel run only as "
                "a disagreement detector. Any pair whose label changed during relabeling is removed from "
                "every profile instead of being assigned the new label."
            ),
            "prompts_used": (
                "Initial labels use the same Phase 1/2 entity-matcher JSON prompt as three-phase v2. "
                "The agent_precision batch prompt is used only to identify unstable pairs; its changed "
                "labels are not adopted."
            ),
            "prompt_source": (
                "Initial prompt: scripts/labeling/run_simple_labeling.py (_label_pair). "
                "Disagreement signal: scripts/experiments/evidence_first_abstain/prompts/"
                "agent_precision_system_prompt.txt via "
                "scripts/labeling/build_drop_changed_profiles.py"
            ),
            "notes": "Conservative noise-removal variant: drop disputed examples rather than flip their labels.",
        },
        "Three-Phase v2 + Closure Bridge Drop": {
            "method": "Three-Phase v2 + Closure Bridge Drop",
            "labeling_model": "gpt-5.2 labels, no extra labeling model",
            "how_it_works": (
                "Starts from the original three-phase v2 labels, builds a graph over positive matches, "
                "identifies positive bridge edges, and removes those bridge pairs from every profile. "
                "All remaining labels stay unchanged."
            ),
            "prompts_used": (
                "No new LLM prompt. The method only filters labels that were already produced by the "
                "three-phase v2 entity-matcher prompt."
            ),
            "prompt_source": "scripts/labeling/build_closure_bridge_profiles.py",
            "notes": "Closure-only noise-removal variant; it can remove many positives on dense datasets.",
        },
        "Three-Phase v2 + Closure Bridge + Relabel-Changed Drop": {
            "method": "Three-Phase v2 + Closure Bridge + Relabel-Changed Drop",
            "labeling_model": "gpt-5.2 -> gpt-5-mini signal, labels unchanged",
            "how_it_works": (
                "Combines two filters with AND logic: a pair is removed only when it is a positive "
                "bridge edge in the original three-phase v2 graph and the batch-relabel pass changed "
                "that pair's label. All remaining labels stay unchanged."
            ),
            "prompts_used": (
                "Initial labels use the same Phase 1/2 entity-matcher JSON prompt as three-phase v2. "
                "The agent_precision batch prompt is used only as a relabel-changed signal; changed "
                "labels are not adopted."
            ),
            "prompt_source": (
                "Initial prompt: scripts/labeling/run_simple_labeling.py (_label_pair). "
                "Bridge filter: scripts/labeling/build_closure_bridge_profiles.py with "
                "--require-relabel-changed. Relabel signal: scripts/experiments/"
                "evidence_first_abstain/prompts/agent_precision_system_prompt.txt"
            ),
            "notes": "Conservative intersection variant: drops only structurally suspicious positives that also changed under relabeling.",
        },
        "Three-Phase v2 + Closure Bridge OR Relabel-Changed Drop": {
            "method": "Three-Phase v2 + Closure Bridge OR Relabel-Changed Drop",
            "labeling_model": "gpt-5.2 -> gpt-5-mini signal, labels unchanged",
            "how_it_works": (
                "Combines the two filters with OR logic: a pair is removed when it is either a "
                "positive bridge edge in the original three-phase v2 graph or the batch-relabel pass "
                "changed that pair's label. All remaining labels stay unchanged."
            ),
            "prompts_used": (
                "Initial labels use the same Phase 1/2 entity-matcher JSON prompt as three-phase v2. "
                "The agent_precision batch prompt is used only as a changed-label signal; changed "
                "labels are not adopted."
            ),
            "prompt_source": (
                "Initial prompt: scripts/labeling/run_simple_labeling.py (_label_pair). "
                "Bridge/OR filter: scripts/labeling/build_closure_bridge_profiles.py with "
                "--include-relabel-changed. Relabel signal: scripts/experiments/"
                "evidence_first_abstain/prompts/agent_precision_system_prompt.txt"
            ),
            "notes": "Aggressive union variant: drops all bridge positives plus every relabel-disagreement pair.",
        },
    }

    rows = [method_rows[name] for name in methods_in_overview if name in method_rows]
    return pd.DataFrame(rows)


def _write_workbook(path: Path, df: pd.DataFrame, method_guide_df: pd.DataFrame) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Overview", index=False)
        method_guide_df.to_excel(writer, sheet_name="Method Guide", index=False)

    workbook = load_workbook(path)
    sheet = workbook["Overview"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    positive_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    negative_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    baseline_fill = PatternFill(fill_type="solid", fgColor="F3F3F3")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in sheet.iter_rows(min_row=2):
        method_cell = row[1]
        f1_cell = row[3]
        delta_cell = row[4]
        if method_cell.value == "Baseline":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = baseline_fill
        if isinstance(f1_cell.value, (int, float)):
            f1_cell.number_format = "0.0%"
        if isinstance(delta_cell.value, (int, float)):
            delta_cell.number_format = "+0.0%;-0.0%;0.0%"
            if delta_cell.value > 0:
                delta_cell.fill = positive_fill
            elif delta_cell.value < 0:
                delta_cell.fill = negative_fill

    for idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = 12
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    guide = workbook["Method Guide"]
    guide.freeze_panes = "A2"
    guide.auto_filter.ref = guide.dimensions
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in guide.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        for cell in row:
            cell.alignment = wrap_alignment

    guide_widths = {
        "A": 30,
        "B": 22,
        "C": 56,
        "D": 85,
        "E": 75,
        "F": 42,
    }
    for column, width in guide_widths.items():
        guide.column_dimensions[column].width = width

    workbook.save(path)


def main() -> None:
    args = _parse_args()
    export_dir = Path(args.export_dir).expanduser().resolve()
    baseline_export_dir = Path(args.baseline_export_dir).expanduser().resolve()
    active_learning_dir = Path(args.active_learning_dir).expanduser().resolve()

    if not export_dir.exists():
        raise FileNotFoundError(f"Export directory does not exist: {export_dir}")
    if not baseline_export_dir.exists():
        raise FileNotFoundError(f"Baseline export directory does not exist: {baseline_export_dir}")

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else export_dir / DEFAULT_OUTPUT_NAME
    )

    baseline_index = _collect_baselines(baseline_export_dir)
    export_best_df = _collect_export_family_best(export_dir, baseline_index)
    active_learning_df = _collect_active_learning_v1(active_learning_dir, baseline_index)
    dataset_frames = _build_dataset_frames(baseline_index, export_best_df, active_learning_df)
    one_sheet_df = _build_one_sheet_rows(dataset_frames)
    methods_in_overview = [
        str(value)
        for value in one_sheet_df["method"].dropna().drop_duplicates().tolist()
    ]
    method_guide_df = _build_method_guide_rows(methods_in_overview)
    _write_workbook(output_path, one_sheet_df, method_guide_df)
    print(output_path)


if __name__ == "__main__":
    main()
